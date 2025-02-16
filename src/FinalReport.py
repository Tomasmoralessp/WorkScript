import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime as dt
import webbrowser

class FinalReport:
    def __init__(self, processed_data):
        """
        processed_data: Diccionario con los DataFrames ya procesados.
        Formato esperado:
            'arrivals': 'df_arrivals',
            'departures': 'df_departures'.
            'roomreport': 'df_roomreport'.
        """

        self.df_arrivals = processed_data.get('arrivals', pd.DataFrame())
        self.df_departures = processed_data.get('departures', pd.DataFrame())
        self.df_roomreport = processed_data.get('roomreport', pd.DataFrame())

        self.df_room_changes = None
        self.df_staysover = None

    def guess_status_logic(self):
        df_coincidences = pd.merge(self.df_arrivals, self.df_roomreport, on=['name'], how='inner', suffixes=('_arr','_room'))

        # Si hay coincidencia y se queda en la misma habitación es "YA AQUÍ"
        self.df_stayovers = pd.merge(self.df_arrivals, self.df_roomreport, on=['name', 'room_number'], how='inner')

        # Si hay coincidencia pero se quedan en otra habitación es un cambio
        self.df_room_changes = df_coincidences[~df_coincidences['name'].isin(self.df_stayovers['name'])]

        # Agregar las columnas 'room_from' y 'room_to'
        self.df_room_changes = self.df_room_changes.rename(columns={'room_number_arr': 'room_from', 'room_number_room': 'room_to'})
        self.df_room_changes = self.df_room_changes[['name', 'room_from', 'room_to', 'arr_date', 'dep_date']]

        # Agregar la columna change y mostrar las columnas
        self.df_room_changes['change'] = list(zip(self.df_room_changes['room_from'],self.df_room_changes['room_to']))
        self.df_room_changes = self.df_room_changes[['name','change','arr_date','dep_date']]


        # Si no hay coincidencia, es una nueva entrada
        self.df_arrivals = self.df_arrivals[~self.df_arrivals['name'].isin(df_coincidences['name'])]

        # Salidas
        self.df_departures = self.df_departures[~self.df_departures['name'].isin(self.df_stayovers['name'])]

    def generate_excel(self):
        # Sustituir por listas tomadas del dataframe
        arrivals = self.df_arrivals['room_number'].tolist()
        departures = self.df_departures['room_number'].tolist()
        changes = self.df_room_changes['change'].tolist()

        # Load the Excel Workbook
        book = load_workbook('data/daily_excel.xlsx')
        sheet = book.active

        # Get current_date
        current_date = dt.today()
        current_date = (current_date.strftime('%A %d/%m/%Y')).upper()


        # Get the number of arrivals, departures and changes
        n_arrivals = len(arrivals)
        n_departures = len(departures)
        n_changes = len(changes)

        # Set the row ranges for each category
        arrivals_row_range = sheet['B8':f'C{ 7+ n_arrivals}']
        departures_row_range = sheet['F8':f'G{7 + n_departures}']
        change_from_row_range = sheet['B34':f'C{33 + n_changes}']
        change_to_row_range = sheet['F34': f'G{33 + n_changes}']

        # Define the colors
        arrivals_fill = PatternFill(start_color="C1F0C8", fill_type="solid")  
        departures_fill = PatternFill(start_color="F2CEEF", fill_type="solid")  
        changes_fill = PatternFill(start_color="CAEDFB", fill_type="solid") 


        # Set the current date
        sheet['B4'] = current_date 


        # Write the arrivals in the excel document
        for index, row in enumerate(arrivals_row_range):
            cell = row[0]
            cell.value = arrivals[index]
            cell.fill = arrivals_fill 
            
        # Write the departures in the excel document
        for index, row in enumerate(departures_row_range):
            cell = row[0]
            cell.value = departures[index]
            cell.fill = departures_fill 

        # Write the room the client is changing from 
        for index, row in enumerate(change_from_row_range):
            cell = row[0]
            cell.value = changes[index][0]
            cell.fill = changes_fill


        # Write the room the client is changing to
        for index, row in enumerate(change_to_row_range):
            cell = row[0]
            cell.value = changes[index][1]
            cell.fill = changes_fill

        book.save('todays_report.xlsx')


    def generate_cards(self):

        html = """ <!DOCTYPE html>
            <html lang="en">
            <head>
            <link href="https://unpkg.com/tailwindcss@^1.0/dist/tailwind.min.css" rel="stylesheet">
            <meta charset="UTF-8">
            <script defer src="https://pyscript.net/latest/pyscript.js"></script>
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Client Cards</title>
            </head>

            <body class="h-100 grid grid-cols-2 gap-8 p-8"> 
            """


        for _, arrival in self.df_arrivals.iterrows():  # ✅ Usamos .iterrows()
            notes = arrival["notes"]  # ✅ Acceso más limpio a notes
            change_info = notes.get("Change")  # ✅ Evitar error si "Change" no existe

            html += f"""
                <div class="bg-white shadow-md rounded-lg overflow-hidden w-full max-w-sm mx-auto h-[220px] flex flex-col">
                <div class="p-3 flex-grow overflow-auto">
                    <div class="flex justify-between items-start mb-2">
                        <div>
                            <div class="flex items-center gap-2 mb-1">
                                <span class="text-xl font-bold text-gray-800">{arrival["room_number"]}</span>
                                <span class="text-base font-medium text-gray-700 truncate">{arrival["name"]}</span>
                                 <div class="ml-2 inline-flex items-center border-l border-r border-gray-300 px-4 h-6">
                                    <span class="text-xs text-gray-500"> </span>
                                </div>
                            </div>
                            <div class="text-xs font-medium text-gray-600 bg-gray-100 px-2 py-0.5 rounded inline-block">
                                {arrival["guest_type"]}
                            </div>
                        </div>
                        <div class="text-right text-xs font-medium text-gray-600">
                            <p>
                                {arrival["arr_date"][:-5]} - {arrival["dep_date"][:-5]}  <!-- ✅ Solo día/mes -->
                            </p>
                        </div>
                    </div>
                    <div class="grid grid-cols-2 gap-1 text-xs mb-2">
                        {f'<p><span class="font-semibold"> Cobrar: </span> {str(arrival["am_to_pay"])} €</p>' if arrival.get("am_to_pay") else ''}

                        {"<p><span class='font-semibold'>Llegada:</span> " + notes.get('Hour_Of_Arrival', '') + "</p>" if notes.get('Hour_Of_Arrival') else ''}
                        
                        {("<p><span class='font-semibold'>Equipo:</span> " + 
                        (" Cuna" if notes.get('Baby_Cot') else '') +
                        (" &" if notes.get('Baby_Cot') and notes.get('High_Chair') else '') +
                        (" Trona" if notes.get('High_Chair') else '') +
                        "</p>") if notes.get('Baby_Cot') or notes.get('High_Chair') else ''}

                        {("<p> <span class='font-semibold'>Ropa Cama:</span> " + 
                        (" Extra Toppers" if notes.get('Toppers') else '') +  
                        (" &" if notes.get('Toppers') and notes.get('Blankets') else '') +
                        (" Blankets" if notes.get('Blankets') else '') +
                        "</p>" ) if notes.get('Toppers') or notes.get('Blankets') else ''}

                        {f'<p><span class="font-semibold"> Guest of Owner</span> ' if notes.get("Guest_Of_Owner") else ''} 

                        {("<p> <span class='font-semibold'> Changes to: " + 
                        f"{change_info['to']} ({change_info['date'][:-5]})" + "</p>") if change_info else ""}
                    </div>
                </div>
                {("<div class='bg-gray-50 p-2 border-t border-gray-200 text-xs text-gray-600 italic overflow-auto max-h-[60px]'>" + notes.get('Others', '') + "</div>") if notes.get('Others') else ''}
            </div>
            """

            html += """
            </body>
            </html>
            """


            # Escribir el contenido HTML a un archivo HTML
            file_path = "client_cards.html"

            with open(file_path, "w", encoding="utf-8") as file:
                file.write(html)

            webbrowser.open(file_path)



