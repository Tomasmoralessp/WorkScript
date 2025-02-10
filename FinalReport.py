# Constantes para idear la lógica 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime as dt

arrivals = [
    {
        "name": "Patricia Hogan",
        "room_number": 118,
        "guest_type": "Owner usage",
        "arr_date": "01/02/2025",
        "dep_date": "08/02/2025",
        "notes": {
            "Baby_Cot": False,
            "High_Chair": False,
            "Hour_Of_Arrival": "14:00",
            "Guest_Of_Owner": True,
            "Toppers": False,
            "Blankets": True,
            "Others": ""
        },
        "am_to_pay": 0.00
    },
    {
        "name": "Hans Berre",
        "room_number": 200,
        "guest_type": "Transient",
        "arr_date": "01/02/2025",
        "dep_date": "05/02/2025",
        "notes": {
            "Baby_Cot": True,
            "High_Chair": True,
            "Hour_Of_Arrival": "21:00",
            "Guest_Of_Owner": False,
            "Toppers": False,
            "Blankets": False,
            "Others": ""
        },
        "am_to_pay": 850.00
    },
    {
        "name": "Sonja Kessler",
        "room_number": 203,
        "guest_type": "Transient",
        "arr_date": "01/02/2025",
        "dep_date": "04/02/2025",
        "notes": {
            "Baby_Cot": False,
            "High_Chair": False,
            "Hour_Of_Arrival": "13:30",
            "Guest_Of_Owner": False,
            "Toppers": False,
            "Blankets": False,
            "Others": "Vegan breakfast"
        },
        "am_to_pay": 645.00
    },
    {
        "name": "Johan Nielsen",
        "room_number": 102,
        "guest_type": "Owner usage",
        "arr_date": "01/02/2025",
        "dep_date": "08/02/2025",
        "notes": {
            "Baby_Cot": False,
            "High_Chair": False,
            "Hour_Of_Arrival": None,
            "Guest_Of_Owner": True,
            "Toppers": False,
            "Blankets": False,
            "Others": "No housekeeping needed"
        },
        "am_to_pay": 0.00
    },
    {
        "name": "Alice Montague",
        "room_number": 105,
        "guest_type": "Transient",
        "arr_date": "01/02/2025",
        "dep_date": "07/02/2025",
        "notes": {
            "Baby_Cot": False,
            "High_Chair": False,
            "Hour_Of_Arrival": None,
            "Guest_Of_Owner": False,
            "Toppers": False,
            "Blankets": False,
            "Others": "Minibar refill daily, Gluten-free food requested"
        },
        "am_to_pay": 1025.50
    },
    {
        "name": "Kristian Berg",
        "room_number": 101,
        "guest_type": "Transient",
        "arr_date": "01/02/2025",
        "dep_date": "06/02/2025",
        "notes": {
            "Baby_Cot": False,
            "High_Chair": False,
            "Hour_Of_Arrival": None,
            "Guest_Of_Owner": False,
            "Toppers": False,
            "Blankets": False,
            "Others": "Quiet room requested"
        },
        "am_to_pay": 820.00
    },
    {
        "name": "Erika Lunden",
        "room_number": 100,
        "guest_type": "Transient",
        "arr_date": "01/02/2025",
        "dep_date": "05/02/2025",
        "notes": {
            "Baby_Cot": False,
            "High_Chair": False,
            "Hour_Of_Arrival": "12:00",
            "Guest_Of_Owner": False,
            "Toppers": False,
            "Blankets": False,
            "Others": ""
        },
        "am_to_pay": 990.00
    }
]

departures = [
    {"name": "Antonio Jose Garcia", "room_number": 100},  # Salida
    {"name": "Mona Kristin Ovrid", "room_number": 101},  # Salida
    {"name": "Karl Green", "room_number": 208},  # Salida
    {"name": "Solfrid Karlsen", "room_number": 102},  # Salida
    {"name": "Line Schobeck", "room_number": 105},  # Salida
    {"name": "Patricia Hogan", "room_number": 118},  # Salida (aunque sea "ya aquí")
    {"name": "Hans Berre", "room_number": 119},  # Salida (aunque cambie de habitación)
    {"name": "Melea Wallenius", "room_number": 111},  # Salida
]

room_report = [
    {"name": "GARCIA ANTONIO JOSE", "room_number": 100, "arr_date": "29/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "OVRID MONA KRISTIN", "room_number": 101, "arr_date": "28/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "HUNTINGTON JEFFREY", "room_number": 201, "arr_date": "30/01/2025", "dep_date": "05/02/2025"},
    {"name": "DYRSTAD STÅLE", "room_number": 205, "arr_date": "27/01/2025", "dep_date": "06/02/2025"},
    {"name": "SIGURDARDÒTTIR TÒRDIS", "room_number": 207, "arr_date": "29/01/2025", "dep_date": "07/02/2025"},
    {"name": "GREEN KARL", "room_number": 208, "arr_date": "30/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "KARLSEN SOLFRID", "room_number": 102, "arr_date": "28/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "TAYEB FATIHA", "room_number": 103, "arr_date": "30/01/2025", "dep_date": "06/02/2025"},
    {"name": "SCHOBECK LINE", "room_number": 105, "arr_date": "29/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "HOGAN PATRICIA", "room_number": 118, "arr_date": "28/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "BERRE HANS", "room_number": 119, "arr_date": "27/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "JORDET ODRUN", "room_number": 117, "arr_date": "30/01/2025", "dep_date": "07/02/2025"},
    {"name": "WALLENIUS MELEA", "room_number": 111, "arr_date": "29/01/2025", "dep_date": "01/02/2025"},  # Salida
    {"name": "MONTAGUE ALICE", "room_number": 105, "arr_date": "27/01/2025", "dep_date": "01/02/2025"}  # Ahora es un "Ya Aquí"
]


# Create df 
df_arrivals = pd.DataFrame(arrivals)

# Expandir la columna notas en multiples columnas separadas
notes_df = pd.json_normalize(df_arrivals['notes'])

# Unir el data frame original con el Dataframe expandido de notas
df_arrivals = df_arrivals.drop(columns='notes').join(notes_df)


df_departures = pd.DataFrame(departures)
df_room_report = pd.DataFrame(room_report)

# Format the names so they share the same structure and are comparable
df_room_report['name'] = df_room_report['name'].str.split().str[::-1].str.join(' ').str.upper()
df_arrivals['name'] = df_arrivals['name'].str.upper()
df_departures['name'] = df_departures['name'].str.upper()

# Report Class para manejar la lógica ha de recibir todos los data frame,

class FinalReport():
  def __init__(self, df_arrivals, df_departures, df_room_report):
    self.df_arrivals = df_arrivals
    self.departures = df_departures
    self.df_room_report = df_room_report
    self.df_room_changes = None
    self.df_stayovers  = None
  

  # TODO: change_from change_to df_room_changes
  def guess_status_logic(self):
    df_coincidences = pd.merge(self.df_arrivals, self.df_room_report, on=['name'], how='inner')

    # 2. Si hay coincidencia y se queda en la misma habitación son un YA AQUI
    self.df_staysover = pd.merge(self.df_arrivals, self.df_room_report, on=['name','room_number'], how='inner')


    # 3. Si hay coincidencia pero se quedan en otra habitación son un cambio
    self.df_room_changes = df_coincidences[~df_coincidences['name'].isin(self.df_staysover['name'])]


    # 4. Si no hay coincidencia son una entrada normal
    self.df_arrivals = self.df_arrivals[~self.df_arrivals['name'].isin(df_coincidences['name'])]

    # 5. Salida
    self.df_departures = self.df_departures[~self.df_departures['name'].isin(self.df_staysover['name'])]

    
 
  def generate_excel(self):
    pass
    


  def generate_cards(self):
    pass


# Sustituir por listas tomadas del dataframe
arrivals = [101, 200, 201, 204, 205, 200]
departures = [204, 207, 203, 100, 104]
changes = [(200, 300), (500, 400), (300, 210), (120, 121), (125, 126)]

# Load the Excel Workbook
book = load_workbook('daily_excel.xlsx')
sheet = book.active

# Get current_date
current_date = dt.today()
current_date = (current_date.strftime('%A %d/%m/%Y')).upper()


# Get the number of arrivals, departures and changes
n_arrivals = len(arrivals)
n_departures = len(departures)
n_changes = len(changes)

# Set the row ranges for each category
arrivals_row_range = sheet['C8':f'C{ 7+ n_arrivals}']
departures_row_range = sheet['G8':f'G{7 + n_departures}']
change_from_row_range = sheet['C34':f'C{33 + n_changes}']
change_to_row_range = sheet['G34': f'G{33 + n_changes}']

# Define the colors
arrivals_fill = PatternFill(start_color="C1F0C8", fill_type="solid")  
departures_fill = PatternFill(start_color="F2CEEF", fill_type="solid")  
changes_fill = PatternFill(start_color="CAEDFB", fill_type="solid") 


# Set the current date
sheet['C4'] = current_date 


# Write the arrivals in the excel document
for index, row in enumerate(arrivals_row_range):
    cell = row[0]
    cell.value = arrivals[index]
    cell.fill = arrivals_fill 
    
# Write the departures in the excel document
for index, row in enumerate(departures_row_range):
    cell = row[0]
    cell.value = departures[index]
    cell.fill = departures_fill 

# Write the room the client is changing from 
for index, row in enumerate(change_from_row_range):
    cell = row[0]
    cell.value = changes[index][0]
    cell.fill = changes_fill


# Write the room the client is changing to
for index, row in enumerate(change_to_row_range):
    cell = row[0]
    cell.value = changes[index][1]
    cell.fill = changes_fill

book.save('todays_report.xlsx')